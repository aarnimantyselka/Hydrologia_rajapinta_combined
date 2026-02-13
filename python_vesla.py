import pandas as pd
import numpy as np
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.series import Series 
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.legend import Legend
from datetime import datetime, timedelta, timezone
import os
import Hydrograafi
import sys

## CSV-tiedoston lataaminen ja tallenuskansion määritys

delimiter = ";"
otsikko_rivit = 0
havainto_paikka = "0409910_W"
vakio = 0.01

## Funktio, joka hakee vedenkorkeusdatan rajapinnasta halutulta aikaväliltä ja tietyltä havaintopaikalta. Datan hakeminen tapahtuu paloissa, jotta vältetään API:n rajaamat datamäärät.
#  Datan hakemisen jälkeen se tallennetaan DataFrameen, jossa on sarakkeet Aika, Arvo ja Paikka_Id. 
#  Lisäksi funktio laskee haluttaessa tasokorjauksen ja lisää sen uutena sarakkeena. 
#  Lopuksi funktio palauttaa DataFramen, johon on lisätty Vuosi ja day-sarakkeet, jotka helpottavat myöhempää pivotointia.
def vedenkorkeus_API(alkuaika, loppuaika, tunnus, taso = None, tasokorjaus_arvo = None):
    # Convert to datetime if needed (optional)
    alkuaika = datetime.strptime(alkuaika, "%Y-%m-%d")
    loppuaika = datetime.strptime(loppuaika, "%Y-%m-%d")

    # Format as string in the required datetime format
    alku_str = alkuaika.strftime("%Y-%m-%dT00:00:00.000Z")
    loppu_str   = loppuaika.strftime("%Y-%m-%dT23:59:59.999Z")  # include full day
    days = (loppuaika.date() - alkuaika.date()).days + 1
    base_url = "https://rajapinnat.ymparisto.fi/api/Hydrologiarajapinta/1.1/odata/Vedenkorkeus?"

    all_rows = []

    # Number of days per query (adjust if needed)
    chunk_days = 499

    current_start = alkuaika

    while current_start <= loppuaika:
        current_end = min(current_start + timedelta(days=chunk_days - 1), loppuaika)

        alku_str = current_start.strftime("%Y-%m-%dT00:00:00.000Z")
        loppu_str = current_end.strftime("%Y-%m-%dT23:59:59.999Z")

        params = {
            "$filter": (
                f"Aika ge datetime'{alku_str}' and "
                f"Aika le datetime'{loppu_str}' and "
                f"Paikka/Nro eq '{tunnus}'"
            )
        }

        response = requests.get(base_url, params=params)
        response.raise_for_status()
        data = response.json()

        rows = data.get("value", [])
        all_rows.extend(rows)

        print(f"Fetched {len(rows)} rows from {current_start.date()} to {current_end.date()}")

        current_start = current_end + timedelta(days=1)

    df_response = pd.DataFrame(all_rows)

    required_cols = {"Paikka_Id", "Aika", "Arvo"}

    if not required_cols.issubset(df_response.columns):
        print("Dataa ei löytynyt")
        return None
    
    df_response = df_response[["Aika","Arvo","Paikka_Id"]]
    df_response['Aika'] = pd.to_datetime(df_response['Aika'])


    paikka_id = df_response['Paikka_Id']
    paikka_id = paikka_id.loc[0]


    if taso == None:
        print("Tasokorjaus tietoa ei saatavilla")
    else:
        df_response[taso] = df_response["Arvo"] + tasokorjaus_arvo
        ## Siirretään korjattu sarake toiseksi
        col_to_move = taso
        cols = df_response.columns.tolist()
        cols.insert(1, cols.pop(cols.index(col_to_move)))

        df_response = df_response[cols]

        print(f"Tasokorjaus {taso} tehty")

    df_response['Vuosi'] = df_response['Aika'].dt.year
    df_response['day'] = df_response['Aika'].dt.dayofyear
    return df_response


def Lumen_aluevesiarvo_API(alkuaika, loppuaika, tunnus):
    # Convert to datetime if needed (optional)
    alkuaika = datetime.strptime(alkuaika, "%Y-%m-%d")
    loppuaika = datetime.strptime(loppuaika, "%Y-%m-%d")

    # Format as string in the required datetime format
    alku_str = alkuaika.strftime("%Y-%m-%dT00:00:00.000Z")
    loppu_str   = loppuaika.strftime("%Y-%m-%dT23:59:59.999Z")  # include full day
    days = (loppuaika.date() - alkuaika.date()).days + 1
    base_url = "https://rajapinnat.ymparisto.fi/api/Hydrologiarajapinta/1.1/odata/LumiAlue?"

    all_rows = []

    # Number of days per query (adjust if needed)
    chunk_days = 499

    current_start = alkuaika

    while current_start <= loppuaika:
        current_end = min(current_start + timedelta(days=chunk_days - 1), loppuaika)

        alku_str = current_start.strftime("%Y-%m-%dT00:00:00.000Z")
        loppu_str = current_end.strftime("%Y-%m-%dT23:59:59.999Z")

        params = {
            "$filter": (
                f"Aika ge datetime'{alku_str}' and "
                f"Aika le datetime'{loppu_str}' and "
                f"Paikka/Nro eq '{tunnus}'"
            )
        }

        response = requests.get(base_url, params=params)
        response.raise_for_status()
        data = response.json()

        rows = data.get("value", [])
        all_rows.extend(rows)

        print(f"Fetched {len(rows)} rows from {current_start.date()} to {current_end.date()}")

        current_start = current_end + timedelta(days=1)

    df_response = pd.DataFrame(all_rows)

    required_cols = {"Paikka_Id", "Aika", "Arvo"}

    if not required_cols.issubset(df_response.columns):
        print("Dataa ei löytynyt")
        return None
    
    df_response = df_response[["Aika","Arvo","Paikka_Id"]]
    df_response['Aika'] = pd.to_datetime(df_response['Aika'])

    return df_response

## Apufunktio, joka hakee viimeisin havainto -päivämäärän jokaiselle havaintopaikalle ja päivittää sen DataFrameen. Funktio hakee datan rajapinnasta vain niille havaintopaikoille, joiden viimeisin havainto ei ole tänään, jotta vältetään turhat API-kutsut.
def viimeisin_havainto_päivitys(df):
    base_url = "https://rajapinnat.ymparisto.fi/api/Hydrologiarajapinta/1.1/odata/Vedenkorkeus?"
    all_rows = []

    for index, row in df.iterrows():
        tila = row.get("Tila")

        if tila == "Lopetettu":
                    print(f"Rivi {index}: lopetettu asema, ohitetaan")
                    continue
        paikka_id = row.get("paikka_id")

        # --- Parse existing ViimeisinHavainto safely ---
        try:
            viimeisin_havainto = datetime.strptime(
                str(row["ViimeisinHavainto"]),
                "%d.%m.%Y"
            )
        except (ValueError, TypeError):
            # If missing or malformed, set default date to 1.1.2026
            viimeisin_havainto = datetime(2026, 1, 1)

        alku_str = viimeisin_havainto.strftime("%Y-%m-%dT23:59:59.999Z")
        current_time_str = (
                datetime.now(timezone.utc)
                .strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
        )

        # --- Skip if latest observation is already today ---
        if viimeisin_havainto.date() == datetime.now(timezone.utc).date():
            print(f"Rivi {index}: ViimeisinHavainto on jo tänään -> ohitetaan")
            continue

        params = {
            "$filter": (
                f"Aika ge datetime'{alku_str}' and "
                f"Aika le datetime'{current_time_str}' and "
                f"Paikka_Id eq {paikka_id}"
            )
        }

        try:
            response = requests.get(base_url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as e:
            print(f"Rivi {index}: API-virhe ({paikka_id}) -> {e}")
            continue

        rows = data.get("value")

        if not rows:
            print(f"Rivi {index}: ei uusia havaintoja")
            continue

        # --- Find latest observation ---
        try:
            latest_row = max(rows, key=lambda r: r["Aika"])
            latest_dt = datetime.fromisoformat(
                latest_row["Aika"].replace("Z", "")
            )
        except (KeyError, ValueError, TypeError):
            print(f"Rivi {index}: virhe API-datassa")
            continue

        # --- Update DataFrame ---
        df.loc[index, "ViimeisinHavainto"] = latest_dt.strftime("%d.%m.%Y")
        print(f"{df.loc[index, 'ViimeisinHavainto']} - {row.get('Nimi')}")
    return df

## Funktio, joka pivotoi DataFramen siten, että rivit ovat päivä-kuukausi yhdistelmiä (mm-dd), 
#  sarakkeet ovat vuosia ja solujen arvot ovat vedenkorkeuksia. 
#  Lisäksi funktio laskee min, max, keskiarvo ja mediaani jokaiselle päivälle ja tallentaa ne erillisiin sarakkeisiin. 
#  Lopuksi funktio laskee min, keskiarvo ja max jokaiselle vuodelle ja tallentaa ne alas uusina riveinä.
def pivot_ja_tallennus(dataframe, havainto_paikka, vakio):

    df = dataframe
    
    df.iloc[:, 1] = df.iloc[:,1] * vakio

    ## muunnetaan pvm sarake pd.datetime sarakkeeksi

    df['date'] = pd.to_datetime(df['Aika'], dayfirst = True, errors= 'coerce')

    ## poimitaan päivämäärästä päivä ja kuukausi sekä vuosi

    df['day'] = df['date'].dt.day
    df['month'] = df['date'].dt.month
    df['day-month'] = '2000-' + df['month'].astype(str).str.zfill(2) + "-" + df['day'].astype(str).str.zfill(2)
    df['year'] = df['date'].dt.year

    # Pivot: rows = day_month, columns = year, values = value
    pivot_df = df.pivot(index='day-month', columns='year', values=havainto_paikka)

    # Sort index to maintain calendar order
    pivot_df.index = pd.to_datetime(pivot_df.index)
    pivot_df = pivot_df.sort_index()
    pivot_df.index = pivot_df.index.strftime('%-d.%-m.')

    # Reset index to make 'Date' a column
    pivot_df.reset_index(inplace=True)
    pivot_df.rename(columns={'day-month': 'Date'}, inplace=True)
    pivot_df['Date'] = pd.to_datetime(pivot_df['Date']).dt.date

    # Lasketaan min, max, keskiarvo ja mediaani vuoden jokaiselle päivälle ja tallennetaan ne erillisiin sarakkeisiin
    value_cols = pivot_df.columns[1:]   # or explicitly list year columns
    df = pivot_df.copy()

    df['Min'] = df[value_cols].min(axis=1, skipna=True)
    df['Max'] = df[value_cols].max(axis=1, skipna=True)
    df['keskiarvo'] = df[value_cols].mean(axis=1, skipna=True)
    df['Mediaani'] = df[value_cols].median(axis=1, skipna=True)

    # Lasketaan Min, keskiarvo ja Max jokaiselle vuodelle ja tallennetaan alas uusina riveinä

    vuosi_sarakkeet = value_cols
    print("ok2")
    min_rivi = ['Min'] + [df[year].min(skipna=True) for year in vuosi_sarakkeet] + [None] * 4
    avg_rivi = ['Avg'] + [df[year].mean(skipna=True) for year in vuosi_sarakkeet] + [None] * 4
    max_rivi = ['Max'] + [df[year].max(skipna=True) for year in vuosi_sarakkeet] + [None] * 4

    ## luodaan uusi dataframe äsken luoduista listoista
    print("ok3")
    summary_df = pd.DataFrame([min_rivi, avg_rivi, max_rivi], columns=df.columns)

    # Lisätään se varsinaiseen dataframeen

    df = pd.concat([df, summary_df], ignore_index=True)

    return df

## Funktio joka laskee prosenttipisteet jokaiselle päivälle ja tallentaa ne uutena sarakkeena.
#  Prosenttipisteet lasketaan nearest rank -menetelmällä, joka on yksinkertainen ja helposti ymmärrettävä tapa laskea prosenttipisteitä.
#  Prosenttipisteet lasketaan vain niille sarakkeille, jotka edustavat vuosia, jotta vältetään virheelliset laskut muille sarakkeille.
def prosentti_pisteet(data_frame):
    prosentit = [5, 10, 25, 50, 75, 90, 95]
    sarake_nimet = [f"P{p}" for p in prosentit]
    # Explicitly select numeric columns (assumed to be year columns)
    vuosi_sarakkeet = get_year_columns_p(data_frame)

    # --- Laske prosenttipisteet ja lisää ne df:ään uusina sarakkeina ---
    for p in prosentit:
        data_frame[f'P{p}'] = data_frame[vuosi_sarakkeet].apply(
            lambda row: nearest_rank_percentile(row, p), axis=1
        )

    return data_frame

## Apufunktio, joka palauttaa DataFramen sarakkeet, jotka edustavat vuosia, vaikka ne olisivatkin merkkijonoja kuten '2000'.
def get_year_columns_p(df):
    """Return columns that represent years, even if they are strings like '2000'."""
    year_cols = []
    for col in df.columns:
        try:
            # Try converting column name to int
            year = int(col)
            if 1900 <= year <= 2100:  # sanity check for valid year
                year_cols.append(col)
        except (ValueError, TypeError):
            continue
    return year_cols

## Funktio joka laskee nearest rank -prosenttipisteen yhdelle riville. Nearest rank -menetelmässä prosenttipisteen sijainti määritetään järjestämällä data ja valitsemalla tietty arvo järjestyksestä. Tämä menetelmä on yksinkertainen ja helposti ymmärrettävä, mutta se ei välttämättä ole yhtä tarkka kuin interpolointimenetelmät, erityisesti pienillä datamäärillä.
def nearest_rank_percentile(arvot, p):
    data = np.sort(pd.to_numeric(arvot, errors="coerce").dropna())
    n = len(data)
    if n==0:
        return np.nan
    rank = int(np.floor((p / 100) * n) + 1)
    if rank < 1:
        rank = 1
    if rank > n:
        rank = n
    return data[rank - 1]  # Python 0-indeksoitu


## Funktio, joka laskee pysyvyysprosentit tietyille kynnysarvoille ja tallentaa ne DataFrameen. 
#  Pysyvyysprosentti kertoo, kuinka suuri osa havainnoista ylittää tietyn kynnysarvon. 
def pysyvyydet(df, vuosi_sarakkeet, step):

    print(vuosi_sarakkeet)
    # 1. Kerätään kaikki arvot yhteen taulukkoon
    kaikki_arvot = df[vuosi_sarakkeet].to_numpy().flatten()
    # 2. Poistetaan puuttuvat arvot
    arvot = kaikki_arvot[~np.isnan(kaikki_arvot)]

    if len(arvot) == 0:
        raise ValueError("No valid data found in the selected columns")

    # Determine min, max, step
    fs_min = arvot.min()
    fs_max = arvot.max()
    fs_väli = step

    thresholds = np.arange(fs_min, fs_max + fs_väli, fs_väli)

    # Calculate exceedance %
    prosentit = []
    for t in thresholds:
        pct = 100 * np.sum(arvot > t) / len(arvot)
        prosentit.append(pct)

    # Construct result DataFrame
    result = pd.DataFrame({
        '%': prosentit[::-1],       
        'Yläraja': thresholds[::-1]
    })

    # Ensure last row is exactly 100% / min
    if result['%'].iloc[-1] != 100:
        result.loc[len(result)] = [100, fs_min]

    # Keep 2 decimals for readability
    result['%'] = result['%'].round(2)
    result['Yläraja'] = result['Yläraja'].round(2)

    # 6. DataFrame tuloksena
    return result
    

## Funktio, joka luo viivakaavion Excel-tiedostoon käyttäen openpyxl-kirjastoa. Kaavio luodaan LineChart-objektina, johon lisätään dataa Reference-objektien avulla.
#  Kaavion ulkoasua muokataan asettamalla otsikko, akselien nimet, ruudukko ja legenda. Lopuksi kaavio lisätään Excel-tiedostoon ja tiedosto tallennetaan.
def linechart(excel_path, otsikko, y_axis, num):

    # Aukaistaan Excel tiedosto ja välilehti

    wb = load_workbook(excel_path)
    ws = wb.active

    chart = LineChart()
    chart.title = otsikko

    chart.y_axis.title = y_axis
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    ## Valitaan visualisoitava data

    max_row = ws.max_row
    max_col = ws.max_column

    # Lisätään kaikki vuodet graafiin

    data = Reference(ws, min_col = 2, min_row=1, max_col=max_col-num, max_row=max_row-3)
    categories = Reference(ws, min_col=1, min_row=2,max_row=max_row-3)

    chart.add_data(data, titles_from_data=True)


    # Lisätään min, max ja mediaani graafiin

    summary_styles = {
        'Min': {'color': '0000FF', 'style': 'sysDot'},
        'Max': {'color': 'FF0000', 'style': 'sysDot'},
        'Mediaani': {'color': '808080', 'style': 'sysDot'}
    }

    for label, props in summary_styles.items():
        for col in range(1, max_col + 1):
            if ws.cell(row=1, column=col).value == label:
                summary_data = Reference(ws, min_col=col, min_row=1, max_row=max_row-3)
                chart.add_data(summary_data, titles_from_data=True)
                s = chart.series[-1]
                s.graphicalProperties.line.solidFill = props['color']
                s.graphicalProperties.line.dashStyle = props['style']
                break

    chart.set_categories(categories)

    # Enable major gridlines
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()

    # Määritellään graafin koko

    chart.width = 20
    chart.height = 15

    ## Määritellään legendin sijainti

    chart.legend.position = "b"
    chart.legend.overlay = False

    ws.add_chart(chart, "M2")
    wb.save(excel_path)
    print("Viivakaavio lisätty Excel-tiedostoon")

## Apufunktio, joka hakee resurssitiedoston absoluuttisen polun, jotta se toimii sekä kehitysympäristössä että PyInstallerilla pakattuna. 
#  Funktio tarkistaa ensin, onko _MEIPASS-attribuutti olemassa, mikä tarkoittaa, että ohjelma on pakattu PyInstallerilla. 
#  Jos se on olemassa, se käyttää sitä pohjana resurssitiedoston polulle. Muussa tapauksessa se käyttää nykyisen työskentelykansion absoluuttista polkua.
def get_resource_path(relative_path):
    """ Get absolute path to resource, works for dev and PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def main():
    pass


if __name__ == "__main__":
    main()












