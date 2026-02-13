import pandas as pd
import numpy as np
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.series import Series 
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.legend import Legend
from datetime import datetime, timedelta
import os
import Hydrograafi_virtaama


## CSV-tiedoston lataaminen ja tallenuskansion määritys

file_path = "C:/Users/A033696/OneDrive - ELY, TET, KEHA/Python_vesla/Maavesi_W_2010_2020.csv"
output_path = "C:/Users/A033696/OneDrive - ELY, TET, KEHA/Python_vesla/Maavesi_2010_2020_W_transformed.xlsx"
delimiter = ";"
otsikko_rivit = 0
havainto_paikka = "0409910_W"
vakio = 0.01

def vedenkorkeus_API(alkuaika, loppuaika, tunnus):
    # Convert to datetime if needed (optional)
    alkuaika = datetime.strptime(alkuaika, "%Y-%m-%d")
    loppuaika = datetime.strptime(loppuaika, "%Y-%m-%d")

    # Format as string in the required datetime format
    alku_str = alkuaika.strftime("%Y-%m-%dT00:00:00.000Z")
    loppu_str   = loppuaika.strftime("%Y-%m-%dT23:59:59.999Z")  # include full day
    days = (loppuaika.date() - alkuaika.date()).days + 1
    base_url = "https://rajapinnat.ymparisto.fi/api/Hydrologiarajapinta/1.1/odata/Virtaama?"

    all_rows = []

    # Number of days per query (adjust if needed)
    chunk_days = 499

    current_start = alkuaika

    while current_start <= loppuaika:
        current_end = min(current_start + timedelta(days=chunk_days - 1), loppuaika)

        alku_str = current_start.strftime("%Y-%m-%dT00:00:00.000Z")
        loppu_str = current_end.strftime("%Y-%m-%dT23:59:59.999Z")

        params = {
            "$filter": (
                f"Aika ge datetime'{alku_str}' and "
                f"Aika le datetime'{loppu_str}' and "
                f"Paikka/Nro eq '{tunnus}'"
            )
        }

        response = requests.get(base_url, params=params)
        response.raise_for_status()
        data = response.json()

        rows = data.get("value", [])
        all_rows.extend(rows)

        print(f"Fetched {len(rows)} rows from {current_start.date()} to {current_end.date()}")

        current_start = current_end + timedelta(days=1)

    df_response = pd.DataFrame(all_rows)

    required_cols = {"Paikka_Id", "Aika", "Arvo"}

    if not required_cols.issubset(df_response.columns):
        print("Dataa ei löytynyt")
        return None
    
    df_response = df_response[["Aika","Arvo","Paikka_Id"]]
    df_response['Aika'] = pd.to_datetime(df_response['Aika'])

    df_response['Vuosi'] = df_response['Aika'].dt.year
    df_response['day'] = df_response['Aika'].dt.dayofyear
    df_response['Arvo'] = df_response['Arvo'].astype(float)
    return df_response


def pivot_ja_tallennus(dataframe, havainto_paikka, vakio = None):

    df = dataframe
    df.iloc[:, 1] = df.iloc[:,1]

    ## muunnetaan pvm sarake pd.datetime sarakkeeksi

    df['date'] = pd.to_datetime(df['Aika'], dayfirst = True, errors= 'coerce')

    ## poimitaan päivämäärästä päivä ja kuukausi sekä vuosi

    df['day'] = df['date'].dt.day
    df['month'] = df['date'].dt.month
    df['day-month'] = '2000-' + df['month'].astype(str).str.zfill(2) + "-" + df['day'].astype(str).str.zfill(2)
    df['year'] = df['date'].dt.year

    # Pivot: rows = day_month, columns = year, values = value
    pivot_df = df.pivot(index='day-month', columns='year', values=havainto_paikka)

    # Sort index to maintain calendar order
    pivot_df.index = pd.to_datetime(pivot_df.index)
    pivot_df = pivot_df.sort_index()
    pivot_df.index = pivot_df.index.strftime('%-d.%-m.')

    # Reset index to make 'Date' a column
    pivot_df.reset_index(inplace=True)
    pivot_df.rename(columns={'day-month': 'Date'}, inplace=True)
    pivot_df['Date'] = pd.to_datetime(pivot_df['Date']).dt.date
    print(pivot_df.head())

    # Lasketaan min, max, keskiarvo ja mediaani vuoden jokaiselle päivälle ja tallennetaan ne erillisiin sarakkeisiin
    df = pivot_df

    df['Min']     = df.iloc[:, 1:].min(axis=1, skipna=True)
    df['Max']     = df.iloc[:, 1:].max(axis=1, skipna=True)
    df['keskiarvo'] = df.iloc[:, 1:].mean(axis=1, skipna=True)
    df['Mediaani']  = df.iloc[:, 1:].median(axis=1, skipna=True)

    # Lasketaan Min, keskiarvo ja Max jokaiselle vuodelle ja tallennetaan alas uusina riveinä

    vuosi_sarakkeet = df.columns[1:-4]

    min_rivi = ['Min'] + [df[year].min(skipna=True) for year in vuosi_sarakkeet] + [None] * 4
    avg_rivi = ['Avg'] + [df[year].mean(skipna=True) for year in vuosi_sarakkeet] + [None] * 4
    max_rivi = ['Max'] + [df[year].max(skipna=True) for year in vuosi_sarakkeet] + [None] * 4

    ## luodaan uusi dataframe äsken luoduista listoista

    summary_df = pd.DataFrame([min_rivi, avg_rivi, max_rivi], columns=df.columns)

    # Lisätään se varsinaiseen dataframeen

    df = pd.concat([df, summary_df], ignore_index=True)

    return df

def prosentti_pisteet(data_frame):
    prosentit = [5, 10, 25, 50, 75, 90, 95]
    sarake_nimet = [f"P{p}" for p in prosentit]
    vuosi_sarakkeet = data_frame.columns[1:-4]


    # --- Laske prosenttipisteet ja lisää ne df:ään uusina sarakkeina ---
    for p in prosentit:
        data_frame[f'P{p}'] = data_frame[vuosi_sarakkeet].apply(
            lambda row: nearest_rank_percentile(row, p), axis=1
        )

    return data_frame


def nearest_rank_percentile(arvot, p):
    data = np.sort(pd.to_numeric(arvot, errors="coerce").dropna())
    n = len(data)
    if n==0:
        return np.nan
    rank = int(np.floor((p / 100) * n) + 1)
    if rank < 1:
        rank = 1
    if rank > n:
        rank = n
    return data[rank - 1]  # Python 0-indeksoitu

def pysyvyydet(dataframe, vuosi_sarakkeet, n_väli):
    kaikki_arvot = dataframe[vuosi_sarakkeet].to_numpy().ravel()
    # 2. Poistetaan puuttuvat arvot
    arvot = kaikki_arvot[~np.isnan(kaikki_arvot)]
    
    fsMin = np.floor(np.min(arvot) * 10) / 10
    fsMax = np.ceil(np.max(arvot) * 10) / 10
    fsVäli = (fsMax - fsMin) / n_väli

    # 4. Väliarvot
    ylärajat = np.arange(fsMin, fsMax + fsVäli, fsVäli)
    
    # Pysyvyysprosentit nousevassa järjestyksessä
    prosentit = [100 * np.sum(arvot <= raja) / len(arvot) for raja in ylärajat]


    # 6. DataFrame tuloksena
    return pd.DataFrame({'%': prosentit, 'Yläraja': ylärajat})
    


def linechart(excel_path, otsikko, y_axis, num):
    
    # Aukaistaan Excel tiedosto ja välilehti

    wb = load_workbook(excel_path)
    ws = wb.active

    chart = LineChart()
    chart.title = otsikko
    
    chart.y_axis.title = y_axis
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    ## Valitaan visualisoitava data

    max_row = ws.max_row
    max_col = ws.max_column

    # Lisätään kaikki vuodet graafiin

    data = Reference(ws, min_col = 2, min_row=1, max_col=max_col-num, max_row=max_row-3)
    categories = Reference(ws, min_col=1, min_row=2,max_row=max_row-3)

    chart.add_data(data, titles_from_data=True)
    

    # Lisätään min, max ja mediaani graafiin

    summary_styles = {
        'Min': {'color': '0000FF', 'style': 'sysDot'},
        'Max': {'color': 'FF0000', 'style': 'sysDot'},
        'Mediaani': {'color': '808080', 'style': 'sysDot'}
    }
    
    for label, props in summary_styles.items():
        for col in range(1, max_col + 1):
            if ws.cell(row=1, column=col).value == label:
                summary_data = Reference(ws, min_col=col, min_row=1, max_row=max_row-3)
                chart.add_data(summary_data, titles_from_data=True)
                s = chart.series[-1]
                s.graphicalProperties.line.solidFill = props['color']
                s.graphicalProperties.line.dashStyle = props['style']
                break
    
    chart.set_categories(categories)

    # Enable major gridlines
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()

    # Määritellään graafin koko

    chart.width = 20
    chart.height = 15

    ## Määritellään legendin sijainti

    chart.legend.position = "b"
    chart.legend.overlay = False

    ws.add_chart(chart, "M2")
    wb.save(excel_path)
    print("Viivakaavio lisätty Excel-tiedostoon")


def viikkotason_voimakkuus(dataframe):
    
    date_col = dataframe.iloc[:,0]
    dataframe = dataframe.iloc[:,1:]

    rolling = dataframe.rolling(window=7, center=True, min_periods=1)
    rolled = rolling.agg(['max', 'min', 'mean'])

    max_df = rolled.xs('max', axis=1, level=1)
    min_df = rolled.xs('min', axis=1, level=1)
    mean_df = rolled.xs('mean', axis=1, level=1)

    voimakkuus_df = (max_df - min_df) / mean_df
    voimakkuus_df.loc[0:2, :] = np.nan
    voimakkuus_df.loc[363:365, :] = np.nan

    summary = pd.DataFrame({
        'min': voimakkuus_df.min(),
        'max': voimakkuus_df.max(),
        'mean': voimakkuus_df.mean()
    }).T  # Transpose so that years are columns

    empty_row = pd.Dataframeempty_row = pd.DataFrame([[np.nan]*len(dataframe.columns)], columns=dataframe.columns)
    intensity_with_summary = pd.concat([voimakkuus_df, empty_row, summary], axis=0, ignore_index=True)

    summary_index = ['min', 'max', 'mean']
    extended_date = pd.concat([date_col, pd.Series([np.nan]), pd.Series(summary_index)], ignore_index=True)

    voimakkuus_df = pd.concat([extended_date.rename('pvm'), intensity_with_summary], axis=1, ignore_index=False)

    return voimakkuus_df

def tulosta_asemat(Paa_vesi_ID, suure):

    base_url = "https://rajapinnat.ymparisto.fi/api/Hydrologiarajapinta/1.1/odata/Paikka?"

    params = {
    "$filter": (
        f"Suure/Nimi eq '{suure}' and "
        f"H_PaaVesal_Id eq {Paa_vesi_ID} "
    )
    }
    response = requests.get(base_url, params=params)
    response.raise_for_status()
    data = response.json()

    df_response = pd.DataFrame(data["value"])

    required_cols = {"Nimi", "Nro", "Paikka_Id"}

    if not required_cols.issubset(df_response.columns):
        print("Dataa ei löytynyt")
        return None
    
    valitut_sarakkeet = df_response[['Nimi', 'Nro', 'Paikka_Id']]
    valitut_sarakkeet.to_excel("paikat.xlsx", index=False)
    os.startfile("paikat.xlsx")

def main():

    alkuaika = "2020-01-01"
    loppuaika = "2025-12-10"
    tunnus = "0405400"
    tulos = vedenkorkeus_API(alkuaika, loppuaika, tunnus)
    output_path = "C:/Users/A033696/OneDrive - ELY, TET, KEHA/Python_vesla/testi.xlsx"
    havainto_paikka = "W N2000"
    ## tulos.to_excel(output_path, index=False)
    ## dataframe = pivot_ja_tallennus(tulos, havainto_paikka, 0.01)
    ##print(dataframe)
    ## Määrittele muuttujien nimet siten kuin ne on csv tiedostoissa määritetty
    muuttuja = "W N2000"

    ## Interaktiivisen kuvaajan teko
    print(tulos)
    app = Hydrograafi_virtaama.init_dash(dataframe=tulos, variable_name = muuttuja)
    app.run(debug=True, use_reloader = False)
    """
    # Kansioiden ja parametrien määrittäminen
    file_path = "C:/Users/A033696/OneDrive - ELY, TET, KEHA/Python_vesla/Maavesi_W_2010_2020.csv"
    output_path = "C:/Users/A033696/OneDrive - ELY, TET, KEHA/Python_vesla/Maavesi_2010_2020_W_transformed.xlsx"
    delimiter = ";"
    otsikko_rivit = 0
    havainto_paikka = "0409800_W"
    vakiokerroin = 0.01
    

    ## Datan pivotointi ja min, max, avg ja mediaanien laskeminen
    dataframe = pivot_ja_tallennus(file_path,delimiter,otsikko_rivit, havainto_paikka, vakiokerroin)
    vuosi_sarakkeet = dataframe.columns[0:-4]
    vuosi_sarakkeet_df = dataframe[vuosi_sarakkeet]
    vain_vuosidata_df = vuosi_sarakkeet_df.iloc[:-3]

    ## Prosenttipisteiden laskemien
    dataframe = prosentti_pisteet(data_frame=dataframe)

    # Pivotoidun datan tallentaminen Exceliin
    dataframe.to_excel(output_path, index=False)
    print("Excel tiedosto luotu onnistuneesti")

    ## Pysyvyyksien laskeminen
    dataframe2 = pysyvyydet(dataframe, vuosi_sarakkeet=vuosi_sarakkeet, n_väli=20)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        dataframe2.to_excel(writer, sheet_name='Pysyvyys', index=False)
        print("Pysyvyydet lisätty")

    ## Viivakaavion lisääminen 
    linechart(output_path, otsikko="Päivittäiset vedenkorkeus havainnot", y_axis= "vedenkorkeus")

    ##Viikkotason voimakkuuden (säännöstelyn) laskeminen

    viikkotason_voimakkuudet_df = viikkotason_voimakkuus(vain_vuosidata_df)

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        viikkotason_voimakkuudet_df.to_excel(writer, sheet_name='vt voimakkuus', index=False)
        print("voimakkuudet lisätty")

    """
# Run the main function
if __name__ == "__main__":
    main()












